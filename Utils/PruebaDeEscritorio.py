from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

class PruebaDeEscritorio:
    __nombre_archivo: str
    __wb: Workbook
    __ws: Worksheet
    __fila: int = 1
    __cabeceras: list[str]
    
    def __init__(self, nombre_archivo: str, cabeceras: list[str]):
        self.__wb = Workbook()
        ws = self.__wb.active
        if ws is None:
            raise Exception("No se pudo crear la hoja de cálculo.")
        else:
            self.__ws = ws
        self.__cabeceras = cabeceras
        for cabecera in cabeceras:
            self.__ws.cell(row=self.__fila, column=cabeceras.index(cabecera) + 1, value=cabecera)
        self.__fila += 1
        self.__wb.save(nombre_archivo)
        self.__nombre_archivo = nombre_archivo
    
    def modificar_variable(self, nombre_variable: str, valor: str):
        self.__ws.cell(row=self.__fila, column=self.__cabeceras.index(nombre_variable) + 1, value=valor)
        self.__fila += 1
        self.__wb.save(self.__nombre_archivo)
    
    def modificar_variables(self, variables: dict[str, str]):
        for nombre_variable, valor in variables.items():
            self.__ws.cell(row=self.__fila, column=self.__cabeceras.index(nombre_variable) + 1, value=valor)
        self.__fila += 1
        self.__wb.save(self.__nombre_archivo)
    
    @property
    def fila(self):
        return self.__fila